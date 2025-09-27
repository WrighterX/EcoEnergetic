import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from sqlalchemy import create_engine
import os
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

engine = create_engine('postgresql+psycopg2://postgres:password@localhost:5432/ecoenergetic')

# Create directories if they don't exist
os.makedirs('charts', exist_ok=True)
os.makedirs('exports', exist_ok=True)

# Function to execute SQL and get DataFrame
def get_df_from_sql(sql):
    return pd.read_sql(sql, engine)

# 1. Pie Chart: Number of charging stations per operator
sql_pie = """
SELECT st.operator_name, COUNT(DISTINCT st.station_id) AS station_count
FROM charging_stations st
JOIN districts d ON st.district_name = d.district_name
JOIN customers c ON st.income_tier = c.income_tier
GROUP BY st.operator_name
"""
df_pie = get_df_from_sql(sql_pie)
fig, ax = plt.subplots()
df_pie.set_index('operator_name')['station_count'].plot.pie(autopct='%1.1f%%', ax=ax)
ax.set_title('Distribution of Charging Stations by Operator')
ax.set_ylabel('')
plt.savefig('charts/pie_stations_by_operator.png')
plt.close()
print(f"Pie chart created: {len(df_pie)} rows, shows distribution of charging stations by operator.")

# 2. Bar Chart: Average charging cost per car model
sql_bar = """
SELECT c.car_model, AVG(cs.total_cost) AS avg_cost
FROM charging_sessions cs
JOIN customers c ON cs.customer_id = c.customer_id
JOIN charging_stations st ON cs.station_id = st.station_id
GROUP BY c.car_model
"""
df_bar = get_df_from_sql(sql_bar)
fig, ax = plt.subplots(figsize=(12, 6))
df_bar.plot.bar(x='car_model', y='avg_cost', ax=ax)
ax.set_title('Average Charging Cost per Car Model')
ax.set_xlabel('Car Model')
ax.set_ylabel('Average Cost')
plt.xticks(rotation=45, ha='right')
plt.tight_layout()
plt.savefig('charts/bar_avg_cost_per_model.png')
plt.close()
print(f"Bar chart created: {len(df_bar)} rows, shows average charging cost per car model.")

# 3. Horizontal Bar Chart: Number of charging sessions per district
sql_hbar = """
SELECT d.district_name, COUNT(cs.session_id) AS session_count
FROM charging_sessions cs
JOIN charging_stations st ON cs.station_id = st.station_id
JOIN districts d ON st.district_name = d.district_name
GROUP BY d.district_name
"""
df_hbar = get_df_from_sql(sql_hbar)
fig, ax = plt.subplots()
df_hbar.plot.barh(x='district_name', y='session_count', ax=ax)
ax.set_title('Number of Charging Sessions per District')
ax.set_xlabel('Session Count')
ax.set_ylabel('District Name')
plt.savefig('charts/hbar_sessions_per_district.png')
plt.close()
print(f"Horizontal bar chart created: {len(df_hbar)} rows, shows number of charging sessions per district.")

# 4. Line Chart: Daily total revenue trend
sql_line = """
SELECT DATE(cs.session_start_time) AS day, SUM(cs.total_cost) AS total_revenue
FROM charging_sessions cs
JOIN charging_stations st ON cs.station_id = st.station_id
JOIN districts d ON st.district_name = d.district_name
GROUP BY day
ORDER BY day
"""
df_line = get_df_from_sql(sql_line)
fig, ax = plt.subplots()
df_line.plot.line(x='day', y='total_revenue', ax=ax)
ax.set_title('Daily Total Revenue Trend')
ax.set_xlabel('Date')
ax.set_ylabel('Total Revenue')
plt.savefig('charts/line_daily_revenue.png')
plt.close()
print(f"Line chart created: {len(df_line)} rows, shows daily total revenue trend over time.")

# 5. Histogram: Distribution of kWh charged per session
sql_hist = """
SELECT cs.kwh_charged
FROM charging_sessions cs
JOIN customers c ON cs.customer_id = c.customer_id
JOIN charging_stations st ON cs.station_id = st.station_id
"""
df_hist = get_df_from_sql(sql_hist)
fig, ax = plt.subplots()
df_hist['kwh_charged'].plot.hist(bins=10, ax=ax)
ax.set_title('Histogram of kWh Charged per Session')
ax.set_xlabel('kWh Charged')
ax.set_ylabel('Frequency')
plt.savefig('charts/hist_kwh_charged.png')
plt.close()
print(f"Histogram created: {len(df_hist)} rows, shows distribution of kWh charged per session.")

# 6. Scatter Plot: Battery capacity vs kWh charged
sql_scatter = """
SELECT c.battery_capacity_kwh, cs.kwh_charged
FROM charging_sessions cs
JOIN customers c ON cs.customer_id = c.customer_id
JOIN charging_stations st ON cs.station_id = st.station_id
"""
df_scatter = get_df_from_sql(sql_scatter)
fig, ax = plt.subplots()
df_scatter.plot.scatter(x='battery_capacity_kwh', y='kwh_charged', ax=ax)
ax.set_title('Scatter Plot of Battery Capacity vs kWh Charged')
ax.set_xlabel('Battery Capacity (kWh)')
ax.set_ylabel('kWh Charged')
plt.savefig('charts/scatter_battery_vs_kwh.png')
plt.close()
print(f"Scatter plot created: {len(df_scatter)} rows, shows relationship between battery capacity and kWh charged.")

# Interactive Plotly chart with time slider, sorted by date
sql_plotly = """
SELECT DATE(cs.session_start_time) AS day, cs.kwh_charged, cs.total_cost, c.car_model
FROM charging_sessions cs
JOIN customers c ON cs.customer_id = c.customer_id
JOIN charging_stations st ON cs.station_id = st.station_id
"""
df_plotly = get_df_from_sql(sql_plotly)
df_plotly = df_plotly.sort_values('day')
fig = px.scatter(df_plotly, x='kwh_charged', y='total_cost',
                 animation_frame='day',
                 color='car_model',
                 size='kwh_charged')
fig.update_layout(title='Interactive Scatter: kWh Charged vs Total Cost by Day')
fig.show()

def export_to_excel(dataframes_dict, filename):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Load workbook for formatting
    wb = load_workbook(filename)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        for col in range(2, ws.max_column + 1):
            col_letter = chr(64 + col)
            rule = ColorScaleRule(start_type='min', start_color='FFAA0000',
                                  mid_type='percentile', mid_value=50, mid_color='FFFFFF00',
                                  end_type='max', end_color='FF00AA00')
            ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)
    wb.save(filename)

sql_export = """
SELECT cs.session_id, cs.customer_id, c.car_model, cs.kwh_charged, cs.total_cost,
       st.district_name, d.income_tier, st.operator_name
FROM charging_sessions cs
JOIN customers c ON cs.customer_id = c.customer_id
JOIN charging_stations st ON cs.station_id = st.station_id
JOIN districts d ON st.district_name = d.district_name
LIMIT 1000
"""
df_export = get_df_from_sql(sql_export)
dataframes_dict = {'Charging_Sessions_Details': df_export}
filename = 'exports/ecoenergetic_report.xlsx'
export_to_excel(dataframes_dict, filename)
total_rows = sum(len(df) for df in dataframes_dict.values())
print(f"Created file {filename}, {len(dataframes_dict)} sheets, {total_rows} rows")